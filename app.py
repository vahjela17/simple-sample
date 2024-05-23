from flask import Flask, jsonify, request,  send_file
from flask_cors import CORS
from sqlalchemy import create_engine, text
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
import logging
import time
from datetime import datetime, timedelta
import pandas as pd
import os
from openpyxl import load_workbook

app = Flask(__name__)
CORS(app)

# Database connection
db_path = 'sqlite:///products.db'
engine = create_engine(db_path, echo=True)
pricing_excel_path = "C:/Users/mikev/Downloads/Wine_Pricing.xlsx"

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def setup_driver():
    """Set up and return a Selenium WebDriver."""
    logging.info("Setting up the driver")
    options = Options()
    driver = webdriver.Chrome(options=options)
    return driver

def login(driver, url_home):
    logging.info("Logging in")
    driver.get(url_home + "user/sign_in")
    WebDriverWait(driver, 12).until(EC.presence_of_element_located((By.XPATH, "//input[@name='user[email]']")))
    input_username = driver.find_element(By.ID, "user_email")
    input_username.send_keys("luke.vahjen@tryondist.com")
    input_password = driver.find_element(By.ID, "user_password")
    input_password.send_keys("qwaug818TY'")
    input_password.submit()

def navigate_to_sample_sheet(driver, url_home):
    """Navigate to a specific page."""
    logging.info("Navigating...")
    tomorrow = (datetime.today() + timedelta(days=1)).strftime('%m-%d-%Y')
    url_with_date = f"{url_home}/samples?sample_list%5Bdate%5D={tomorrow}"
    driver.get(url_with_date)
    driver.get(url_home + "samples/4196107")

def load_pricing_data(pricing_excel_path):
    """Load pricing data from Excel and return a dictionary of Product ID to Unit Pricing."""
    df = pd.read_excel(pricing_excel_path, sheet_name='05-16-2024 Pricing')
    
    # Ensure ProductID is treated as a string to handle mixed data types
    df['ProductID'] = df['ProductID'].astype(str)
    
    # Strip any non-numeric characters from 'Unit Price' and convert to numeric
    df['Unit Price'] = pd.to_numeric(df['Unit Price'].replace(r'[\$,]', '', regex=True))
    
    # Filter out rows where 'ProductID' is not numeric
    df = df[df['ProductID'].str.isnumeric()]
    
    # Convert ProductID back to integer if necessary
    df['ProductID'] = df['ProductID'].astype(int)
    
    pricing_dict = df.set_index('ProductID')['Unit Price'].to_dict()
    return pricing_dict

def insert_values_from_data(driver, pricing_dict, product_ids):
    """Insert values from the provided data into web form."""
    wait = WebDriverWait(driver, 10)
    for product_id in product_ids:
        try:
            input_field = wait.until(EC.presence_of_element_located((By.ID, "vendor_product_version_input")))
            input_field.clear()
            input_field.send_keys(product_id)
            input_field.send_keys(Keys.ARROW_DOWN)
            time.sleep(1)  # Wait for the dropdown to appear
            input_field.send_keys(Keys.RETURN)
            time.sleep(1)  # Adjust timing based on page behavior
            
            # Ensure product_id is int for pricing_dict lookup
            unit_price = str(pricing_dict.get(int(product_id)))
            
            # Find the corresponding modal trigger and click it
            modal_trigger_xpath = f"//a[@data-toggle='modal'][@data-target='#list-entry-prices-form-{product_id}']"
            modal_trigger = wait.until(EC.element_to_be_clickable((By.XPATH, modal_trigger_xpath)))
            modal_trigger.click()
            time.sleep(1)  # Adjust timing based on page behavior
            
            #open_modal_and_insert_values(driver, unit_price)
        except TimeoutException:
            logging.error(f"Timeout while interacting with input field for product ID: {product_id}")
        except Exception as e:
            logging.error(f"Error interacting with input field for product ID: {product_id} - {e}")
'''
def open_modal_and_insert_values(driver, unit_price):
    """Open the modal popup and insert values."""
    try:
        wait = WebDriverWait(driver, 10)
        modal = wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "modal-content")))

        # Insert "Your Price" into the label input field
        input_label = WebDriverWait(modal, 10).until(EC.presence_of_element_located((By.ID, "list_entry_list_entry_prices_attributes_0_label_display")))
        input_label.clear()
        input_label.send_keys("Your Price")

        # Insert the unit price into the price input field
        input_price = WebDriverWait(modal, 10).until(EC.presence_of_element_located((By.ID, "list_entry_list_entry_prices_attributes_0_price_per_unit")))
        input_price.clear()
        input_price.send_keys(unit_price)
        input_price.send_keys(Keys.RETURN)
        time.sleep(1)  # Adjust timing based on page behavior

        # Close the modal
        close_button = modal.find_element(By.XPATH, "//button[@data-dismiss='modal']")
        close_button.click()
        time.sleep(1)  # Adjust timing based on page behavior

    except TimeoutException:
        logging.error("Timeout while waiting for modal to appear or interact with modal elements.")
    except Exception as e:
        logging.error(f"Error interacting with modal: {e}")
'''
def generate_PDF(driver):
    wait = WebDriverWait(driver, 10)
    generate_pdf_link = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='content']/div/div[1]/div[3]/div/div[2]/div[1]/div[1]/a[3]")))
    generate_pdf_link.click()

def wait_for_download_link(driver, timeout=30):
    try:
        wait = WebDriverWait(driver, timeout)
        link = wait.until(EC.presence_of_element_located((By.XPATH, "//a[contains(@href, 'downloads') and text()='Click here']")))
        return link.get_attribute('href')
    except TimeoutException:
        logging.error("Timeout waiting for the download link to appear")
        return None

@app.route('/generate_form', methods=['POST'])
def generate_form():
    data = request.get_json()
    account = data.get('account')
    reason = data.get('reason')
    selectedProducts = data.get('selectedProducts')

    # Load the spreadsheet using openpyxl
    file_path = 'sample_form.xlsx'
    wb = load_workbook(file_path)
    ws = wb.active

    # Clear the necessary cells
    ws['G4'] = ''  # G4
    ws['G6'] = ''  # G6
    ws['B8'] = ''  # B8
    ws['B9'] = ''  # B9
    for row in range(13, ws.max_row + 1):
        ws[f'B{row}'] = ''
        ws[f'D{row}'] = ''

    # Populate the cells
    today = datetime.today().strftime('%m-%d-%Y')
    tomorrow = (datetime.today() + timedelta(days=1)).strftime('%m-%d-%Y')
    ws['G4'] = today  # G4
    ws['G6'] = tomorrow  # G6
    ws['B8'] = account  # B8
    ws['B9'] = reason  # B9

    for i, product in enumerate(selectedProducts, start=13):
        ws[f'B{i}'] = product['productId']  # Column B
        ws[f'D{i}'] = product['productName']  # Column D

    # Save the updated spreadsheet in a specific directory
    output_dir = os.path.join(app.root_path, 'generated_files')
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, f"{today}-sample-request.xlsx")
    wb.save(output_file)

    # Return the link to download the file
    return jsonify({"excel_link": f"/download/{today}-sample-request.xlsx"})

@app.route('/download/<filename>', methods=['GET'])
def download_file(filename):
    file_path = os.path.join(app.root_path, 'generated_files', filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        return jsonify({"error": "File not found"}), 404
    
@app.route('/search', methods=['GET'])
def search():
    query = request.args.get('q', '').lower()
    phrases = query.split()
    with engine.connect() as conn:
        sql = """
            SELECT "ProductID", "ProductName", "UnitPrice"
            FROM products
            WHERE {}
        """.format(' AND '.join(['LOWER("ProductName") LIKE :phrase{}'.format(i) for i in range(len(phrases))]))
        
        params = {'phrase{}'.format(i): f"%{phrase}%" for i, phrase in enumerate(phrases)}
        result = conn.execute(text(sql), params)
        products = [{'ProductID': row[0], 'ProductName': row[1], 'UnitPrice': row[2]} for row in result]
    return jsonify(products)

@app.route('/process', methods=['POST'])
def process():
    data = request.json
    product_ids = [product['productId'] for product in data]
    logging.info(f"Processing product IDs: {product_ids}")

    # Run the Selenium script with the provided data
    pdf_link = run_selenium_script(product_ids)
    return jsonify({"message": "Products processed successfully", "pdf_link": pdf_link})

def run_selenium_script(product_ids):
    url_home = "https://www.sevenfifty.com/"
    driver = setup_driver()
    try:
        login(driver, url_home)
        navigate_to_sample_sheet(driver, url_home)
        pricing_dict = load_pricing_data(pricing_excel_path)
        insert_values_from_data(driver, pricing_dict, product_ids)
        generate_PDF(driver)
        pdf_link = wait_for_download_link(driver)
        return pdf_link
    except Exception as e:
        logging.error(f"An error occurred: {e}")
        return None
    finally:
        driver.quit()

if __name__ == '__main__':
    app.run(debug=True)
